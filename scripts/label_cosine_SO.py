import matplotlib
import numpy as np
from gensim.models import KeyedVectors
from openpyxl import load_workbook
import fasttext as ft
from sklearn import metrics
import seaborn as sns
import matplotlib.pyplot as plt

wb2 = load_workbook('DatasetsLabels.xlsx')
model = 'SO'
fastext = KeyedVectors.load_word2vec_format('SO_vectors_200.bin', binary=True)

def clean(term):
    a = str(term).strip('$').replace("-", ' ').lower().split(" ")
    print(a)
    return a

def get_embeddings(terms):
    embeddings = []
    for term in terms:
        words = clean(term)
        t = []
        for x in words:
            if x in fastext:
                t.append(fastext.get_vector(x))
            else:
                t.append(np.random.uniform(low=-1, high=1, size=(200,)))

        embedding = np.mean(t, axis=0)
        embeddings.append(list(embedding))

    return embeddings


font = {'family': 'normal',
        'size': 6}

matplotlib.rc('font', **font)

raw_data = []
raw_labl = []
raw_tupl = []
for name in wb2.sheetnames:
    terms = []
    ws = wb2[name]
    print(ws.max_row, ws.min_row)
    for row in ws.iter_rows(min_row=ws.min_row, max_col=1, max_row=ws.max_row):
        for cell in row:
            if cell.value:
                terms.append(cell.value)

    embeddings = get_embeddings(terms)
    similarities = metrics.pairwise.cosine_similarity(embeddings)
    iterate_indices = np.tril_indices(similarities.shape[0], k=-1)
    flat_sims = list(similarities[iterate_indices])
    raw_data.extend(flat_sims)
    raw_labl.extend([name]*len(flat_sims))
    tupl = []
    for i, j in zip(*iterate_indices):
        s = 0
        tupl.append([terms[i], terms[j]])
    raw_tupl.extend(tupl)
    sns.heatmap(similarities, yticklabels=terms, xticklabels=terms)
    plt.savefig(f'similarities_{name}.pdf', format='pdf', dpi=1200, bbox_inches='tight')
    plt.clf()

import csv
with open(f'raw_label_similarities_{model}.csv', 'wt') as outf:
    writer = csv.writer(outf)
    writer.writerow(["dataset", "similarity",'label a', 'label b'])
    for lbl, d, (x,y) in zip(raw_labl, raw_data, raw_tupl):
        writer.writerow([lbl, d, x, y])
