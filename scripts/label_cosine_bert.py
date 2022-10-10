import matplotlib
import numpy as np
from openpyxl import load_workbook
from transformers import AutoTokenizer, AutoModel
import torch
from sklearn import metrics
import seaborn as sns
import matplotlib.pyplot as plt

wb2 = load_workbook('DatasetsLabels.xlsx')

tokenizer = AutoTokenizer.from_pretrained("microsoft/codebert-base")
model = AutoModel.from_pretrained("microsoft/codebert-base")


def get_embeddings(terms):
    embeddings = []
    for term in terms:
        nl_tokens = tokenizer.tokenize(term)
        tokens = [tokenizer.cls_token] + [tokenizer.sep_token]
        tokens_ids = tokenizer.convert_tokens_to_ids(tokens)
        context_embeddings = model(torch.tensor(tokens_ids)[None, :])[0]
        sentence_embedding = torch.mean(context_embeddings, dim=1).squeeze()
        embeddings.append(list(sentence_embedding.detach().numpy()))

    return embeddings


font = {'family': 'normal',
        'size': 6}

matplotlib.rc('font', **font)

raw_data = []
raw_labl = []

remap = {'MUDABlue': 'MUDABlue', 'LACT': 'LACT', 'Ohasi': 'Ohasi', 'InformatiCup': 'ClassifyHub',
         'Vasquez 2014 API': 'Vasquez', 'Le Clair Paper': 'Le Clair', 'LASCAD': 'LASCAD',
         'Awesome-Java': 'Awesome-Java', 'HiGitClass-AI': 'HiGitClass-AI', 'HiGitClass-BIO': 'HighGitClass-BIO',
         'Ours': 'Ours'}

for name in wb2.sheetnames:
    if name not in remap:
        continue
    terms = []
    ws = wb2[name]
    print(ws.max_row, ws.min_row)
    for row in ws.iter_rows(min_row=ws.min_row, max_col=1, max_row=ws.max_row):
        for cell in row:
            if cell.value:
                terms.append(cell.value)

    embeddings = get_embeddings(terms)
    similarities = metrics.pairwise.cosine_similarity(embeddings)
    iterate_indices = np.tril_indices(similarities.shape[0], k=-1)
    flat_sims = list(similarities[iterate_indices])
    raw_data.extend(flat_sims)
    raw_labl.extend([remap[name]] * len(flat_sims))
    sns.heatmap(similarities, yticklabels=terms, xticklabels=terms)
    plt.savefig(f'similarities_{name}.pdf', format='pdf', dpi=1200, bbox_inches='tight')
    plt.clf()

import csv

with open(f'raw_label_similarities_CodeBERT.csv', 'wt') as outf:
    writer = csv.writer(outf)
    writer.writerow(["dataset", "similarity"])
    for lbl, d in zip(raw_labl, raw_data):
        writer.writerow([lbl, d])
