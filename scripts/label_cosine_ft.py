import matplotlib
import numpy as np
from openpyxl import load_workbook
import fasttext as ft
from sklearn import metrics
import seaborn as sns
import matplotlib.pyplot as plt
import csv

wb2 = load_workbook('DatasetsLabels.xlsx')
model = 'fasttext'
fastext = ft.load_model('/home/sasce/PycharmProjects/ComponentSemantics/data/models/fastText/wiki.en.bin')


def get_embeddings(terms):
    embeddings = []
    for term in terms:
        embedding = fastext.get_sentence_vector(term)
        embeddings.append(embedding)

    return embeddings


font = {'family': 'normal',
        'size': 6}

matplotlib.rc('font', **font)

raw_data = []
raw_labl = []
raw_tupl = []
remap = {'MUDABlue': 'MUDABlue', 'LACT': 'LACT', 'Ohasi': 'Ohasi', 'InformatiCup': 'ClassifyHub',
         'Vasquez 2014 API': 'Vasquez', 'Le Clair Paper': 'Le Clair', 'LASCAD': 'LASCAD',
         'Awesome-Java': 'Awesome-Java', 'HiGitClass-AI': 'HiGitClass-AI', 'HiGitClass-BIO': 'HiGitClass-BIO',
         'Di Sipio': 'Di Sipio', 'Ours': 'Ours', 'Ohashi': 'Ohashi', 'Sharma': 'Sharma', 'Borges': 'Borges'}

for name in wb2.sheetnames:
    if name not in remap:
        continue
    terms = []
    ws = wb2[name]
    print(ws.max_row, ws.min_row)
    for row in ws.iter_rows(min_row=ws.min_row, max_col=1, max_row=ws.max_row):
        for cell in row:
            if cell.value:
                terms.append(cell.value)

    embeddings = get_embeddings(terms)
    similarities = metrics.pairwise.cosine_similarity(embeddings)
    iterate_indices = np.tril_indices(similarities.shape[0], k=-1)
    flat_sims = list(similarities[iterate_indices])
    raw_data.extend(flat_sims)
    raw_labl.extend([remap[name]] * len(flat_sims))
    tupl = []
    for i, j in zip(*iterate_indices):
        s = 0
        tupl.append([terms[i], terms[j]])
    raw_tupl.extend(tupl)
    sns.heatmap(similarities, yticklabels=terms, xticklabels=terms)
    plt.savefig(f'similarities_{remap[name]}_{model}.pdf', format='pdf', dpi=1200, bbox_inches='tight')
    plt.clf()

with open(f'raw_label_similarities_{model}.csv', 'wt') as outf:
    writer = csv.writer(outf)
    writer.writerow(["dataset", "similarity", 'label a', 'label b'])
    for lbl, d, (x, y) in zip(raw_labl, raw_data, raw_tupl):
        writer.writerow([lbl, d, x, y])
