import matplotlib
import numpy as np
from openpyxl import load_workbook
import fasttext as ft
from sklearn import metrics
import seaborn as sns
import matplotlib.pyplot as plt

wb2 = load_workbook('DatasetsLabels.xlsx')

fastext = ft.load_model('/home/sasce/PycharmProjects/ComponentSemantics/data/models/fastText/wiki.en.bin')


def get_embeddings(terms):
    embeddings = []
    for term in terms:
        embedding = fastext.get_sentence_vector(term)
        embeddings.append(embedding)

    return embeddings


font = {'family': 'normal',
        'size': 6}

matplotlib.rc('font', **font)

for name in wb2.sheetnames:
    terms = []
    ws = wb2[name]
    print(ws.max_row, ws.min_row)
    for row in ws.iter_rows(min_row=ws.min_row, max_col=1, max_row=ws.max_row):
        for cell in row:
            if cell.value:
                terms.append(cell.value)

    embeddings = get_embeddings(terms)
    similarities = metrics.pairwise.cosine_similarity(embeddings)
    iterate_indices = np.tril_indices(similarities.shape[0])
    sns.heatmap(similarities, yticklabels=terms, xticklabels=terms)
    plt.savefig(f'similarities_{name}.pdf', format='pdf', dpi=1200, bbox_inches='tight')
    plt.clf()
